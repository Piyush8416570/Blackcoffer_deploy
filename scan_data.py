import streamlit as st
import streamlit.components.v1 as components
def add_dynamic_background():
    components.html(
        """
        <style>
            .stApp {
                background: transparent;
            }
            .stApp > header {
                background-color: transparent !important;
            }
        </style>
        <canvas id="gradient-canvas"></canvas>
        <script>
            const canvas = document.getElementById('gradient-canvas');
            const ctx = canvas.getContext('2d');

            let width, height, gradient;

            function resize() {
                width = window.innerWidth;
                height = window.innerHeight;
                canvas.width = width;
                canvas.height = height;

                gradient = ctx.createLinearGradient(0, 0, width, height);
                gradient.addColorStop(0, '#00C9FF');
                gradient.addColorStop(1, '#92FE9D');
            }

            function animate() {
                ctx.clearRect(0, 0, width, height);

                const time = Date.now() * 0.001;
                const x = Math.sin(time) * 0.5 + 0.5;
                const y = Math.cos(time) * 0.5 + 0.5;

                gradient = ctx.createLinearGradient(width * x, height * y, width * (1 - x), height * (1 - y));
                gradient.addColorStop(0, '#00C9FF');
                gradient.addColorStop(1, '#92FE9D');

                ctx.fillStyle = gradient;
                ctx.fillRect(0, 0, width, height);

                requestAnimationFrame(animate);
            }

            window.addEventListener('resize', resize);
            resize();
            animate();
        </script>
        """,
        height=0,
    )

# st.title("Dynamic Background Streamlit App")
# st.subheader("This app has a background that changes based on cursor movement.")
# st.write("Move your cursor over the page to see the background change color.")


# ------------------------------------ Just for Dynamic in Streamlit APP
import pandas as pd
pd.set_option("styler.render.max_elements", 1500000)  # Set a value slightly higher than your dataframe size

# import streamlit as st
import io
import base64
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from traitlets import default

from generate_LOI import handle_generate_lois
from Export_data import handle_export
from refresh_button import reset_session_state

def process_spreadsheet(df):
    # Remove the dollar sign from the 'Last Sale Price' and 'Loan Balance' columns
    # df['Loan Balance'] = df['Loan Balance'].replace(r'[\$,]', '', regex=True).astype(float)
    # df['Last Sale Price'] = df['Last Sale Price'].replace(r'[\$,]', '', regex=True).astype(float)

    # # Calculate Percent Equity and Equity columns
    # df['Percent Equity'] = (df['Last Sale Price'] - df['Loan Balance']) / df['Last Sale Price'] * 100
    # df['Percent Equity'] = df['Percent Equity'].round(2)
    df['Equity'] = df['Last Sale Price'] - df['Total Loan Balance']

    # Highlight rows with low equity
    low_equity_mask = df['Percent Equity'] < 15

    # Format numeric columns
    df['Total Loan Balance'] = df['Total Loan Balance'].apply(lambda x: f"${x:,.2f}")
    df['Last Sale Price'] = df['Last Sale Price'].apply(lambda x: f"${x:,.2f}")
    df['Percent Equity'] = df['Percent Equity'].apply(lambda x: f"{x:.2f}%")
    df['Equity'] = df['Equity'].apply(lambda x: f"${x:,.2f}")

    return df, low_equity_mask

def style_excel(writer, df, low_equity_mask):
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Apply yellow background to low equity rows in 'Address' column
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    address_col_index = df.columns.get_loc('Property Address') + 1  # openpyxl is 1-indexed
    for row_idx, is_low_equity in enumerate(low_equity_mask, start=2):  # start=2 because Excel is 1-indexed and we have a header row
        if is_low_equity:
            cell = worksheet.cell(row=row_idx, column=address_col_index)
            cell.fill = yellow_fill

    # Set column widths
    for idx, col in enumerate(df.columns):
        column_letter = get_column_letter(idx + 1)
        column_width = max(df[col].astype(str).map(len).max(), len(col))
        worksheet.column_dimensions[column_letter].width = column_width + 2

def get_file_download_link(df, low_equity_mask):
    file_format = st.selectbox("Choose file format:", ["Excel (.xlsx)", "CSV (.csv)"])
    
    if file_format == "Excel (.xlsx)":
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            style_excel(writer, df, low_equity_mask)
        
        output.seek(0)
        b64 = base64.b64encode(output.getvalue()).decode()
        return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="processed_data.xlsx">Download Processed Excel</a>'
    
    elif file_format == "CSV (.csv)":
        csv = df.to_csv(index=False)
        b64 = base64.b64encode(csv.encode()).decode()
        return f'<a href="data:file/csv;base64,{b64}" download="processed_data.csv">Download Processed CSV</a>'


def handle_scan_data(df):
    processed_df, low_equity_mask = process_spreadsheet(df)
    st.write("Processed Data Preview:")
    st.dataframe(processed_df.style.apply(lambda x: ['background-color: green' if v else '' for v in low_equity_mask], subset=['Property Address'], axis=0))
    
    # Statistics
    st.write(f"Total rows: {len(df)}")
    st.write(f"Rows with low equity (< 15%): {low_equity_mask.sum()}")

    return processed_df, low_equity_mask


def main():
    add_dynamic_background()
    st.title("Contract Creator APP")
    st.subheader("We Buy Houses Anywhere LLC")
    # st.write(" ") Is we wanna to write something that is needed we can write here.

    # Add this at the top of your main app code, before any other UI elements
    if st.button("Refresh"):
        reset_session_state()
        st.rerun() # using rerun instead of experimental_rerun() vcz it is depreciated

    # Initialize session state variables if they don't exist
    if 'use_default_template' not in st.session_state:
        st.session_state.use_default_template = False
    if 'uploaded_template' not in st.session_state:
        st.session_state.uploaded_template = None
    if 'company_name' not in st.session_state:
        st.session_state.company_name = "We Buy Houses Anywhere LLC"
    if 'your_name' not in st.session_state:
        st.session_state.your_name = "Justin Pickell"
    if 'executed' not in st.session_state:
        st.session_state.executed = False

    uploaded_file = st.file_uploader("Choose a CSV or Excel file", type=['csv', 'xlsx'])
    if uploaded_file is not None:
        # Check if the file is a CSV or an Excel file
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
            # Convert 'Apn' column to string and handle NaN values
            df['Apn'] = df['Apn'].fillna('').astype(str)
            st.write("CSV file uploaded successfully!")
        elif uploaded_file.name.endswith('.xlsx'):
            df = pd.read_excel(uploaded_file)
            # Convert 'Apn' column to string and handle NaN values
            df['Apn'] = df['Apn'].fillna('').astype(str)
            st.write("Excel file uploaded successfully!")
        else:
            st.error("Unsupported file format. Please upload a CSV or Excel file.")

        st.write("Original Data Preview:")
        st.dataframe(df.head(10))
        st.write(f"Rows count: {len(df)}")

        # Calculate Percent Equity here
        df['Total Loan Balance'] = df['Total Loan Balance'].replace(r'[\$,]', '', regex=True).astype(float)
        df['Mls Listing Amount'] = df['Mls Listing Amount'].replace(r'[\$,]', '', regex=True).astype(float)
        df['Percent Equity'] = (df['Mls Listing Amount'] - df['Total Loan Balance']) / df['Mls Listing Amount'] * 100
        df['Percent Equity'] = df['Percent Equity'].round(2)
      

        # Sidebar filters
        with st.sidebar:
            st.header("Filters")

            # Add more filters based on the new sheet's columns
            loan_types = st.multiselect("Loan Types", options=df['Loan Type'].unique(), default=None)
            status = st.multiselect("Status", options=df['Lead Status'].unique(), default=None)
            # property_states = st.multiselect("Property State", options=df['Property State'].unique(), default=None)
            bedroom_count = st.slider("Minimum Bedroom Count", min_value=int(df['Bedroom Count'].min()), max_value=int(df['Bedroom Count'].max()), value=(int(df['Bedroom Count'].min()), int(df['Bedroom Count'].max())))
            bathroom_count = st.slider("Minimum Bathroom Count", min_value=int(df['Bathroom Count'].min()), max_value=int(df['Bathroom Count'].max()), value=(int(df['Bathroom Count'].min()), int(df['Bathroom Count'].max())))
            max_equity = st.selectbox("Max Equity (%)", options=[None, 5, 10, 15, 20, 25], index=4)
            max_interest_rate = st.selectbox("Max Loan Est Interest Rate", options=[None, 3, 3.5, 4, 4.5, 5, 5.5, 6, 6.5, 7], index=7)
            owner_occupied = st.selectbox("Owner Occupied", options=["Yes", "No", "Any"], index=2)

            # Apply filters
            filtered_df = df.copy()

            if loan_types:
                filtered_df = filtered_df[filtered_df['Loan Type'].isin(loan_types)]
            if status:
                filtered_df = filtered_df[filtered_df['Lead Status'].isin(status)]
            # if property_states:
                # filtered_df = filtered_df[filtered_df['Property State'].isin(property_states)]
            if bedroom_count:
                filtered_df = filtered_df[(filtered_df['Bedroom Count'] >= bedroom_count[0]) & (filtered_df['Bedroom Count'] <= bedroom_count[1])]
            if bathroom_count:
                filtered_df = filtered_df[(filtered_df['Bathroom Count'] >= bathroom_count[0]) & (filtered_df['Bathroom Count'] <= bathroom_count[1])]
            if max_equity:
                filtered_df = filtered_df[filtered_df['Percent Equity'].replace(r'[%]', '', regex=True).astype(float) <= max_equity]
            if max_interest_rate:
                filtered_df = filtered_df[filtered_df['Loan Est Interest Rate'] <= max_interest_rate]
            if owner_occupied != "Any":
                filtered_df = filtered_df[filtered_df['Owner Occupied'] == (owner_occupied == "Yes")]

        st.write("Filtered Data Preview:")
        st.dataframe(filtered_df.head(10))
        st.write(f"Rows count: {len(filtered_df)}")

        menu_options = ["1. Scan Data", "2. Generate LOIs", "3. Export"]
        choice = st.selectbox("Select an action:", menu_options)

        if st.button("Execute"):
            st.session_state.executed = True

        if st.session_state.executed:
            processed_df, low_equity_mask = handle_scan_data(filtered_df)
            if choice == "1. Scan Data":
                # processed_df, low_equity_mask = handle_scan_data(filtered_df)
                st.markdown(get_file_download_link(processed_df, low_equity_mask), unsafe_allow_html=True)
            elif choice == "2. Generate LOIs":
                handle_generate_lois(processed_df)
            elif choice == "3. Export":
                handle_export(processed_df)
        # if st.button("Execute"):
        #     processed_df, low_equity_mask = handle_scan_data(filtered_df)
        #     if choice == "1. Scan Data":
        #         # processed_df, low_equity_mask = handle_scan_data(filtered_df)
        #         st.markdown(get_excel_download_link(processed_df, low_equity_mask), unsafe_allow_html=True)
        #     elif choice == "2. Generate LOIs":
        #         handle_generate_lois(processed_df)
        #     elif choice == "3. Export":
        #         handle_export(filtered_df)

if __name__ == "__main__":
    main()
