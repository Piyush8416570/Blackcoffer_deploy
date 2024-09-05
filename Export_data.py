
import streamlit as st
import pandas as pd
import io
import base64
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os

def handle_export(df):
    st.write("Exporting data...")
    
    # Get the list of generated LOI files
    loi_files = [f for f in os.listdir() if f.startswith("LOI_") and f.endswith(".pdf")]
    
    # Create a new column for LOI file names
    df['LOI_File'] = ''
    
    # Populate the LOI_File column
    for file in loi_files:
        index = int(file.split('_')[1].split('.')[0])
        if index < len(df):
            df.at[index, 'LOI_File'] = file
    
    # Create an Excel file
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Merged Data"
    
    # Write the DataFrame to the Excel sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    
    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Save the workbook to the BytesIO object
    workbook.save(output)
    output.seek(0)
    
    # Create a download link for the Excel file
    b64 = base64.b64encode(output.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="merged_data.xlsx">Download Merged Excel File</a>'
    st.markdown(href, unsafe_allow_html=True)

    st.success("Data exported successfully!")